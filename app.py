#!/usr/bin/env python3
"""agent_shubham — a tiny local replica of the ghostwriter ingest + run flow.

What it does (all against your LOCAL stack, no container worker of its own):
  * lists recordings stored in MinIO (recordings/<name>/...) and whether each has a
    recorded_flows DB row,
  * uploads a pasted Playwright .py + params JSON (a flat {key: value} param set): builds the sibling params workbook,
    stores both in MinIO, and upserts the recorded_flows row (same shape ghostwriter writes),
  * runs a recording by shelling out to your LOCAL aetherion CLI
    (`./.venv/bin/aetherion agent 'ACT Agent' '<payload>' --wait`), so the job is picked up
    by the agent/worker you run in your own terminal — not a packaged container.

Run it with the act_agent venv (which already has fastapi/uvicorn/boto3/openpyxl/psycopg2):
    cd act-v2 && act_agent/.venv/bin/python agent_shubham/app.py
then open http://localhost:8765
"""
from __future__ import annotations

import ast
import csv
import io
import json
import os
import re
import socket
import subprocess
import tempfile
from pathlib import Path
from typing import Any

import boto3
import psycopg2
import uvicorn
from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# --------------------------------------------------------------------------- paths / config
HERE = Path(__file__).resolve().parent
# The runner project dir (the "act" repo / ACT Agent); the TEST_RUNNER_DIR env var keeps its legacy
# name so existing shell configs keep working. It holds the venv + aetherion CLI used below.
TEST_RUNNER_DIR = Path(os.environ.get("TEST_RUNNER_DIR", HERE.parent / "act")).resolve()
AETHERION_BIN = os.environ.get("AETHERION_BIN", str(TEST_RUNNER_DIR / ".venv" / "bin" / "aetherion"))
DOWNLOADS_DIR = TEST_RUNNER_DIR / "downloads"


def _load_dotenv(path: Path) -> dict[str, str]:
    cfg: dict[str, str] = {}
    if path.is_file():
        for line in path.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                cfg[k.strip()] = v.strip().strip('"').strip("'")
    return cfg


_ENV = _load_dotenv(TEST_RUNNER_DIR / ".env")


def _cfg(key: str, default: str = "") -> str:
    return (os.environ.get(key) or _ENV.get(key) or default).strip()


def _build_local_aetherion_cli_env() -> dict[str, str]:
    """Force the trigger CLI to use the local repo config instead of ~/.config/aetherion."""
    env = dict(os.environ)
    for key, value in _ENV.items():
        name = str(key or "").strip().upper()
        if not name or value is None:
            continue
        text = str(value).strip()
        if text:
            env[name] = text

    cli_home = Path(tempfile.gettempdir()) / "agent_shubham_aetherion_home"
    cfg_dir = cli_home / ".config" / "aetherion"
    cfg_dir.mkdir(parents=True, exist_ok=True)

    cfg: dict[str, str] = {}
    target_host = str(env.get("AETHERION_TARGET_HOST") or "").strip()
    if target_host:
        cfg["TARGET_HOST"] = target_host
    namespace = str(env.get("AETHERION_NAMESPACE") or "").strip()
    if namespace:
        cfg["NAMESPACE"] = namespace
    (cfg_dir / "config.json").write_text(json.dumps(cfg, indent=2), encoding="utf-8")
    env["HOME"] = str(cli_home)
    return env


DEFAULT_AFTER_ACTION_WAIT_MS = int(_cfg("DEFAULT_AFTER_ACTION_WAIT_MS", "0") or "0")
DEFAULT_MULTI_LINE_SHEET_NAME = "line_items"
# Fixed column that joins a line-item row to its header row (present in BOTH sheets). Hardcoded,
# not configurable -- every multi-header repeatable recording links its lines to headers by ref_id
# (kept in lockstep with the runner's excel_to_json_parse.MATCH_KEY).
MATCH_KEY = "ref_id"


STORAGE_ENDPOINT = _cfg("STORAGE_ENDPOINT", "http://localhost:9000")
STORAGE_ACCESS_KEY = _cfg("STORAGE_ACCESS_KEY")
STORAGE_SECRET_KEY = _cfg("STORAGE_SECRET_KEY")
# The bucket the LOCAL agent reads: TENANT_ID wins, else STORAGE_ACTIVITIES_BUCKET.
BUCKET = _cfg("TENANT_ID") or _cfg("STORAGE_ACTIVITIES_BUCKET") or "local-dev-bucket"

PG = dict(
    host=_cfg("POSTGRES_HOST", "localhost"),
    port=int(_cfg("POSTGRES_PORT", "5435")),  # aetherion-postgresql is published on host :5435
    user=_cfg("POSTGRES_USER", "aetherion"),
    password=_cfg("POSTGRES_PASSWORD", "aetherion"),
    dbname=_cfg("POSTGRES_DB", "aetherion"),
)
DEFAULT_USER_ID = _cfg("USER_ID", "4562a98e-809c-40e8-bc3c-6426bc5d47aa")


def _s3():
    return boto3.client(
        "s3",
        endpoint_url=STORAGE_ENDPOINT,
        aws_access_key_id=STORAGE_ACCESS_KEY,
        aws_secret_access_key=STORAGE_SECRET_KEY,
        region_name="us-east-1",
    )


def _pg():
    return psycopg2.connect(connect_timeout=4, **PG)


# --------------------------------------------------------------------------- params helpers
def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, tuple)):
        return "|".join(_cell(v) for v in value)
    return str(value)


def _coerce_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "yes", "on"}:
            return True
        if normalized in {"0", "false", "no", "off", ""}:
            return False
    return bool(value)


_KNOWN_MULTI_LINE_KEYS = {
    "line_description": "description",
    "description": "description",
    "item": "item",
    "memo_line": "memo_line",
    "uom": "uom",
    "quantity": "quantity",
    "unit_price": "unit_price",
    "amount": "amount",
    "tax_classification": "tax_classification",
    "transaction_business_category": "transaction_business_category",
    "rule": "rule",
    "type": "type",
    "revenue_period": "revenue_period",
}


# Sheet/wrapper names that must never become parameter keys. "Flow Context" is the
# now-removed flow-context spec sheet; the runner dropped flow context entirely (use
# ai_extract() instead), so any such entry is discarded rather than shipped as a param.
_NON_PARAM_KEYS = {
    "flow context", "flow_context", "flowcontext", "flow io", "flow_io",
    "context io", "context_io", "input output", "input_output", "output_input",
}


def _is_non_param_key(key: Any) -> bool:
    return str(key or "").strip().lower() in _NON_PARAM_KEYS


def _maybe_dict(value: Any) -> dict | None:
    """Return a dict if value is one, or a string repr of one (Python literal or JSON)."""
    if isinstance(value, dict):
        return value
    text = str(value or "").strip()
    if not (text.startswith("{") and text.endswith("}")):
        return None
    for parse in (ast.literal_eval, json.loads):
        try:
            parsed = parse(text)
        except (ValueError, SyntaxError, TypeError):
            continue
        if isinstance(parsed, dict):
            return parsed
    return None


def _flatten_param_set(row: dict[str, Any]) -> dict[str, str]:
    """Normalise one parameter set to a flat {key: value} dict.

    Repairs the corruption where a whole row was nested under a sheet-name key with the
    real params stringified as a Python/JSON dict (e.g. {"Sheet1": "{'app_url': ...}",
    "Flow Context": "{...}"}). Such a wrapper made `{{placeholders}}` unresolvable and the
    run fail. We unwrap any dict-valued cell into its own keys and drop flow-context keys.
    """
    flat: dict[str, str] = {}
    for key, value in row.items():
        if _is_non_param_key(key):
            continue
        nested = _maybe_dict(value)
        if nested is not None:
            for nested_key, nested_value in nested.items():
                if _is_non_param_key(nested_key):
                    continue
                flat[str(nested_key)] = _cell(nested_value)
        else:
            flat[str(key)] = _cell(value)
    return flat


def normalize_param_sets(payload: Any, *, allow_empty: bool = False) -> list[dict[str, str]]:
    if isinstance(payload, dict) and "params" in payload:
        raw = payload.get("params") or []
    elif isinstance(payload, list):
        raw = payload
    elif isinstance(payload, dict):
        raw = [payload]
    else:
        raise ValueError("params must be a dict, list of dicts, or {'params': [...]}")
    out = [_flatten_param_set(entry) for entry in raw if isinstance(entry, dict)]
    out = [row for row in out if row]
    if not out:
        if allow_empty:
            # A recording uploaded straight from the recorder may legitimately have no
            # params/{{placeholders}}; ship a single empty set instead of rejecting it.
            return [{}]
        raise ValueError("no parameter sets found")
    return out


def _multi_line_key_in(payload: Any) -> str | None:
    """Return the repeatable-rows key if the payload uses it, else None."""
    if not isinstance(payload, dict):
        return None
    return DEFAULT_MULTI_LINE_SHEET_NAME if DEFAULT_MULTI_LINE_SHEET_NAME in payload else None


def _normalize_multi_line_rows(payload: Any) -> list[dict[str, str]]:
    if payload is None:
        return []
    key = _multi_line_key_in(payload)
    if key is not None:
        raw = payload.get(key) or []
    elif isinstance(payload, list):
        raw = payload
    else:
        return []
    rows = [_flatten_param_set(entry) for entry in raw if isinstance(entry, dict)]
    return [row for row in rows if row]


def _payload_explicitly_sets_multi_line(payload: Any) -> bool:
    return _multi_line_key_in(payload) is not None


def _normalize_field_key(value: Any) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", str(value or "").strip()).strip("_").lower()


def _normalize_match_key(value: Any) -> str:
    return _normalize_field_key(value)


def _extract_row_value_by_key(row: dict[str, Any], match_key: str) -> str:
    normalized_match_key = _normalize_match_key(match_key)
    if not normalized_match_key or not isinstance(row, dict):
        return ""
    for key, value in row.items():
        if _normalize_field_key(key) != normalized_match_key:
            continue
        return str(value or "").strip()
    return ""


def _payload_without_multi_line(payload: Any) -> Any:
    if _multi_line_key_in(payload) is None:
        return payload
    cleaned = dict(payload)
    cleaned.pop(DEFAULT_MULTI_LINE_SHEET_NAME, None)
    return cleaned


def _normalize_repeatable_block_config(payload: Any) -> dict[str, Any] | None:
    if payload is None:
        return None
    if isinstance(payload, str):
        text = payload.strip()
        if not text:
            return None
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            payload = {"enabled": True, "sheet_name": DEFAULT_MULTI_LINE_SHEET_NAME, "prompt": text}
    if not isinstance(payload, dict):
        return None
    if not _coerce_bool(payload.get("enabled", True), default=True):
        return None
    sheet_name = _safe_name(str(payload.get("sheet_name") or DEFAULT_MULTI_LINE_SHEET_NAME)) or DEFAULT_MULTI_LINE_SHEET_NAME
    prompt = str(payload.get("prompt") or "").strip()
    return {
        "enabled": True,
        "sheet_name": sheet_name,
        "prompt": prompt,
    }


def _normalize_repeatable_blocks_config(
    payload: Any = None,
    *,
    legacy_payload: Any = None,
) -> list[dict[str, Any]]:
    raw = payload if payload is not None else legacy_payload
    if raw is None:
        return []
    if isinstance(raw, dict):
        if "repeatable_blocks" in raw:
            raw = raw.get("repeatable_blocks")
        else:
            raw = [raw]
    elif isinstance(raw, str):
        text = raw.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            raw = [text]
        else:
            raw = parsed.get("repeatable_blocks") if isinstance(parsed, dict) and "repeatable_blocks" in parsed else parsed
    if not isinstance(raw, list):
        return []
    blocks: list[dict[str, Any]] = []
    for item in raw:
        normalized = _normalize_repeatable_block_config(item)
        if normalized:
            blocks.append(normalized)
    return blocks


def _repeatable_blocks_from_recording_config(recording_config: Any) -> list[dict[str, Any]]:
    config = recording_config if isinstance(recording_config, dict) else {}
    blocks = _normalize_repeatable_blocks_config(config.get("repeatable_blocks"))
    if blocks:
        return blocks
    return _normalize_repeatable_blocks_config(legacy_payload=config.get("repeatable_line_items"))


def _resolve_upload_multi_line_rows(
    *,
    payload: Any,
    param_sets: list[dict[str, str]],
    repeatable_blocks: list[dict[str, Any]] | None,
    recording_name: str,
    bucket: str,
    overwrite: bool,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    multi_line_rows = _normalize_multi_line_rows(payload)
    if not repeatable_blocks:
        return param_sets, multi_line_rows

    explicit_multi_line = _payload_explicitly_sets_multi_line(payload)
    if not multi_line_rows:
        param_sets, inferred_multi_line_rows = _infer_multi_line_rows_from_param_sets(param_sets)
        multi_line_rows = inferred_multi_line_rows

    # If the user overwrites an existing repeatable-line recording from the editor
    # and the payload omitted the multi_line block, preserve the saved sheet instead
    # of silently rewriting the workbook with only the params sheet.
    if not multi_line_rows and overwrite and not explicit_multi_line:
        try:
            _existing_params, existing_multi_line_rows, _existing_key = _load_saved_runtime_payload(recording_name, bucket)
        except Exception:
            existing_multi_line_rows = []
        if existing_multi_line_rows:
            multi_line_rows = existing_multi_line_rows

    return param_sets, multi_line_rows


def _infer_multi_line_rows_from_param_sets(param_sets: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    header_rows: list[dict[str, str]] = []
    multi_line_rows: list[dict[str, str]] = []
    for row in param_sets:
        header_row: dict[str, str] = {}
        line_row: dict[str, str] = {}
        for key, value in row.items():
            normalized_key = str(key or "").strip()
            mapped = _KNOWN_MULTI_LINE_KEYS.get(normalized_key)
            if mapped:
                line_row[mapped] = value
            else:
                header_row[normalized_key] = value
        header_rows.append(header_row)
        if line_row:
            multi_line_rows.append(line_row)
    return header_rows, multi_line_rows


def _headers(rows: list[dict[str, str]]) -> list[str]:
    seen, order = set(), []
    for row in rows:
        for k in row:
            if k not in seen:
                seen.add(k)
                order.append(k)
    return order


def build_params_csv(param_sets: list[dict[str, str]]) -> bytes:
    headers = _headers(param_sets)
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(headers)
    for row in param_sets:
        w.writerow([row.get(h, "") for h in headers])
    return buf.getvalue().encode("utf-8")


def build_params_xlsx(
    param_sets: list[dict[str, str]],
    *,
    multi_line_rows: list[dict[str, str]] | None = None,
    multi_line_sheet_name: str = DEFAULT_MULTI_LINE_SHEET_NAME,
) -> bytes:
    import openpyxl

    # Materialize the fixed ref_id join column in BOTH sheets so each header row is explicitly
    # linked to its repeated rows: header rows get 1, 2, 3...; repeated rows keep their own value
    # or default to "1" (the first header).
    match_key = MATCH_KEY

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "params"
    ph = _headers(param_sets)
    if match_key and match_key not in ph:
        ph = [match_key, *ph]
    ws.append(ph)
    for idx, row in enumerate(param_sets, start=1):
        ws.append(
            [str(idx) if h == match_key and match_key not in row else row.get(h, "") for h in ph]
        )
    if multi_line_rows is not None:
        sheet = wb.create_sheet(title=_safe_name(multi_line_sheet_name) or DEFAULT_MULTI_LINE_SHEET_NAME)
        mh = _headers(multi_line_rows)
        if match_key and match_key not in mh:
            mh = [match_key, *mh]
        if mh:
            sheet.append(mh)
            for row in multi_line_rows:
                sheet.append(
                    [
                        row.get(h, "1" if h == match_key else "")
                        for h in mh
                    ]
                )
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _safe_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", (name or "").strip()).strip("._")


def _start_url(script: str, first_params: dict[str, str]) -> str:
    m = re.search(r"page\.goto\(\s*[\"']([^\"']+)[\"']", script)
    if not m:
        return ""
    url = m.group(1)
    ph = re.match(r"^\{\{\s*([\w.-]+)\s*\}\}$", url.strip())
    if ph:
        return first_params.get(ph.group(1), url)
    return url


def _placeholders(script: str) -> list[str]:
    return sorted(set(re.findall(r"\{\{\s*([\w.-]+)\s*\}\}", script)))


def _recording_config_key(name: str) -> str:
    return f"recordings/{name}/{name}_recording_config.json"


def _build_recording_config(
    name: str,
    *,
    prompt: str = "",
    repeatable_blocks: list[dict[str, Any]] | None = None,
) -> dict[str, Any] | None:
    normalized_prompt = str(prompt or "").strip()
    normalized_repeatable_blocks = [dict(block) for block in (repeatable_blocks or []) if isinstance(block, dict)]
    if not normalized_prompt and not normalized_repeatable_blocks:
        return None
    config: dict[str, Any] = {
        "version": 1,
        "recording_name": name,
    }
    if normalized_prompt:
        config["prompt"] = normalized_prompt
    if normalized_repeatable_blocks:
        config["repeatable_blocks"] = normalized_repeatable_blocks
    return config


# --------------------------------------------------------------------------- DB ops
def db_rows_by_name(names: list[str]) -> dict[str, dict[str, Any]]:
    if not names:
        return {}
    try:
        with _pg() as cn, cn.cursor() as cur:
            cur.execute(
                "SELECT name, id, file_name, data_file_name, start_url FROM recorded_flows WHERE name = ANY(%s)",
                (names,),
            )
            return {
                r[0]: {"id": str(r[1]), "file_name": r[2], "data_file_name": r[3], "start_url": r[4]}
                for r in cur.fetchall()
            }
    except Exception:
        return {}


def upsert_recorded_flow(*, name, file_name, data_file_name, start_url, user_id, overwrite) -> dict[str, Any]:
    if overwrite:
        sql = """
            INSERT INTO recorded_flows (id, name, file_name, data_file_name, start_url, created_by, updated_by)
            VALUES (gen_random_uuid(), %(name)s, %(file_name)s, %(data_file_name)s, %(start_url)s, %(uid)s, %(uid)s)
            ON CONFLICT (name) DO UPDATE SET
                file_name = EXCLUDED.file_name,
                data_file_name = EXCLUDED.data_file_name,
                start_url = EXCLUDED.start_url,
                updated_by = EXCLUDED.updated_by
            RETURNING id, (xmax = 0) AS inserted;
        """
    else:
        sql = """
            INSERT INTO recorded_flows (id, name, file_name, data_file_name, start_url, created_by, updated_by)
            VALUES (gen_random_uuid(), %(name)s, %(file_name)s, %(data_file_name)s, %(start_url)s, %(uid)s, %(uid)s)
            ON CONFLICT (name) DO NOTHING
            RETURNING id, true AS inserted;
        """
    params = dict(name=name, file_name=file_name, data_file_name=data_file_name, start_url=start_url, uid=user_id)
    with _pg() as cn, cn.cursor() as cur:
        cur.execute(sql, params)
        row = cur.fetchone()
        cn.commit()
    if row is None:
        return {"id": None, "inserted": False, "conflict": True}
    return {"id": str(row[0]), "inserted": bool(row[1]), "conflict": False}


# --------------------------------------------------------------------------- app
app = FastAPI(title="agent_shubham")
app.mount("/downloads", StaticFiles(directory=str(DOWNLOADS_DIR), check_dir=False), name="downloads")


@app.get("/api/scripts")
def list_scripts(bucket: str = ""):
    bkt = bucket or BUCKET
    s3 = _s3()
    paginator = s3.get_paginator("list_objects_v2")
    folders = []
    for page in paginator.paginate(Bucket=bkt, Prefix="recordings/", Delimiter="/"):
        folders.extend(p["Prefix"] for p in page.get("CommonPrefixes", []))
    items = []
    for folder in folders:
        name = folder[len("recordings/"):].rstrip("/")
        if not name:
            continue
        keys = {o["Key"] for o in s3.list_objects_v2(Bucket=bkt, Prefix=folder).get("Contents", [])}
        # Only count folders that follow the runner's convention: recordings/<name>/<name>.py
        canonical_py = f"{folder}{name}.py"
        if canonical_py not in keys:
            continue
        params = next((f"{folder}{name}{ext}" for ext in ("_params.xlsx", "_params.csv") if f"{folder}{name}{ext}" in keys), "")
        items.append({"name": name, "py_key": canonical_py, "params_key": params})
    db = db_rows_by_name([i["name"] for i in items])
    for i in items:
        i["has_db"] = i["name"] in db
    items.sort(key=lambda x: x["name"].lower())
    return {"bucket": bkt, "scripts": items}


@app.get("/api/script")
def get_script(name: str, bucket: str = ""):
    bkt = bucket or BUCKET
    s3 = _s3()
    safe = _safe_name(name)
    py_key = f"recordings/{safe}/{safe}.py"
    try:
        py_text = s3.get_object(Bucket=bkt, Key=py_key)["Body"].read().decode("utf-8", "replace")
    except Exception as exc:
        raise HTTPException(404, f"script not found: {py_key} ({exc})")
    params_rows, multi_line_rows, params_key = _load_saved_runtime_payload(safe, bkt)
    recording_config = _load_recording_config(safe, bkt)
    db = db_rows_by_name([safe]).get(safe)
    return {"name": safe, "py_key": py_key, "py_text": py_text, "params_key": params_key,
            "params": params_rows, "placeholders": _placeholders(py_text), "db": db,
            "recording_config": recording_config, "line_items": multi_line_rows}


def _parse_params_back(raw: bytes, ext: str) -> list[dict[str, str]]:
    if ext.endswith(".csv"):
        rows = list(csv.reader(io.StringIO(raw.decode("utf-8", "replace"))))
    else:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb["params"] if "params" in wb.sheetnames else wb.active
        rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
        wb.close()
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if len(rows) < 2:
        return []
    headers = [str(h).strip() for h in rows[0]]
    raw_sets = [{headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))} for r in rows[1:]]
    # Flatten so a corrupt workbook (sheet-name headers + stringified-dict cells) reads
    # back as flat params instead of being re-shipped to the agent as {Sheet1: "...", ...}.
    return [row for row in (_flatten_param_set(s) for s in raw_sets) if row]


def _parse_multi_line_back(
    raw: bytes,
    *,
    sheet_name: str = DEFAULT_MULTI_LINE_SHEET_NAME,
) -> list[dict[str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    target = None
    for candidate in wb.sheetnames:
        if _safe_name(candidate) == _safe_name(sheet_name):
            target = wb[candidate]
            break
    if target is None:
        wb.close()
        return []
    rows = [[("" if c is None else str(c)) for c in r] for r in target.iter_rows(values_only=True)]
    wb.close()
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if len(rows) < 2:
        return []
    headers = [_safe_name(str(h).strip()) for h in rows[0]]
    parsed: list[dict[str, str]] = []
    for row in rows[1:]:
        values = {
            headers[i]: row[i]
            for i in range(min(len(headers), len(row)))
            if headers[i] and str(row[i] or "").strip()
        }
        if values:
            parsed.append(values)
    return parsed


def _load_saved_runtime_payload(name: str, bucket: str) -> tuple[list[dict[str, str]], list[dict[str, str]], str]:
    s3 = _s3()
    safe = _safe_name(name)
    params_rows: list[dict[str, str]] = []
    multi_line_rows: list[dict[str, str]] = []
    params_key = ""
    recording_config = _load_recording_config(safe, bucket)
    repeatable_blocks = _repeatable_blocks_from_recording_config(recording_config)
    repeatable = repeatable_blocks[0] if repeatable_blocks else None
    multi_line_sheet_name = (
        _safe_name(str(repeatable.get("sheet_name") or DEFAULT_MULTI_LINE_SHEET_NAME))
        if isinstance(repeatable, dict)
        else DEFAULT_MULTI_LINE_SHEET_NAME
    ) or DEFAULT_MULTI_LINE_SHEET_NAME
    for ext in ("_params.xlsx", "_params.csv"):
        key = f"recordings/{safe}/{safe}{ext}"
        try:
            raw = s3.get_object(Bucket=bucket, Key=key)["Body"].read()
            params_key = key
            params_rows = _parse_params_back(raw, ext)
            if ext.endswith(".xlsx"):
                multi_line_rows = _parse_multi_line_back(raw, sheet_name=multi_line_sheet_name)
            break
        except Exception:
            continue
    return params_rows, multi_line_rows, params_key


def _load_recording_config(name: str, bucket: str) -> dict[str, Any]:
    s3 = _s3()
    safe = _safe_name(name)
    config_key = _recording_config_key(safe)
    try:
        raw = s3.get_object(Bucket=bucket, Key=config_key)["Body"].read()
    except Exception:
        return {}
    try:
        data = json.loads(raw.decode("utf-8", "replace"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


class UploadBody(BaseModel):
    name: str = ""
    script: str
    params: Any  # dict | list | {"params":[...],"context":[...]} | JSON string
    fmt: str = "xlsx"
    overwrite: bool = True
    user_id: str = ""
    bucket: str = ""
    prompt: str = ""
    instance: str = ""
    repeatable_blocks: Any = None


class ParamsXlsxBody(BaseModel):
    name: str = ""
    params: Any


def _resolve_upload_name(raw_name: str, payload: Any) -> str:
    direct = _safe_name(raw_name)
    if direct:
        return direct

    candidates: list[str] = []
    if isinstance(payload, dict):
        for key in ("name", "recording_name", "script_name", "test_suite_id"):
            value = str(payload.get(key) or "").strip()
            if value:
                candidates.append(value)
        raw_params = payload.get("params")
        if isinstance(raw_params, list) and raw_params and isinstance(raw_params[0], dict):
            for key in ("name", "recording_name", "script_name", "test_suite_id"):
                value = str(raw_params[0].get(key) or "").strip()
                if value:
                    candidates.append(value)

    for candidate in candidates:
        safe = _safe_name(candidate)
        if safe:
            return safe
    return ""


@app.post("/api/params-xlsx")
def download_params_xlsx(body: ParamsXlsxBody):
    try:
        payload = json.loads(body.params) if isinstance(body.params, str) else body.params
        param_sets = normalize_param_sets(payload, allow_empty=True)
        multi_line_rows = _normalize_multi_line_rows(payload)
        raw = build_params_xlsx(param_sets, multi_line_rows=multi_line_rows)
    except Exception as exc:
        raise HTTPException(400, f"failed to build params workbook: {exc}")
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    safe = _safe_name(body.name or "recording_name") or "recording_name"
    dest = DOWNLOADS_DIR / f"{safe}_params_preview.xlsx"
    dest.write_bytes(raw)
    return {"ok": True, "download_name": dest.name, "download_url": _report_url_for_dest(dest)}


@app.post("/api/upload")
def upload(body: UploadBody):
    bkt = body.bucket or BUCKET
    payload = json.loads(body.params) if isinstance(body.params, str) else body.params
    name = _resolve_upload_name(body.name, payload)
    if not name:
        raise HTTPException(
            400,
            "recording name is required. Send 'name' in the request body, or include one of "
            "'recording_name', 'script_name', or 'test_suite_id' in the payload.",
        )
    try:
        param_sets = normalize_param_sets(payload, allow_empty=True)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    repeatable_blocks = _normalize_repeatable_blocks_config(body.repeatable_blocks)
    param_sets, multi_line_rows = _resolve_upload_multi_line_rows(
        payload=payload,
        param_sets=param_sets,
        repeatable_blocks=repeatable_blocks,
        recording_name=name,
        bucket=bkt,
        overwrite=body.overwrite,
    )

    fmt = body.fmt if body.fmt in ("xlsx", "csv") else "xlsx"
    if fmt == "csv" and repeatable_blocks:
        raise HTTPException(400, "repeatable blocks require xlsx because csv cannot store a dedicated repeatable sheet.")
    if fmt == "xlsx":
        primary_block = repeatable_blocks[0] if repeatable_blocks else None
        multi_line_sheet_name = (
            str(primary_block.get("sheet_name") or DEFAULT_MULTI_LINE_SHEET_NAME)
            if isinstance(primary_block, dict)
            else DEFAULT_MULTI_LINE_SHEET_NAME
        )
        params_bytes, ext, ct = build_params_xlsx(
            param_sets,
            multi_line_rows=multi_line_rows if multi_line_rows else None,
            multi_line_sheet_name=multi_line_sheet_name,
        ), "_params.xlsx", \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        params_bytes, ext, ct = build_params_csv(param_sets), "_params.csv", "text/csv"

    py_bytes = body.script.encode("utf-8")
    py_key = f"recordings/{name}/{name}.py"
    params_key = f"recordings/{name}/{name}{ext}"
    start_url = _start_url(body.script, param_sets[0])
    missing = [p for p in _placeholders(body.script) if p not in param_sets[0]]
    recording_config = _build_recording_config(
        name,
        prompt=body.prompt,
        repeatable_blocks=repeatable_blocks,
    )
    recording_config_key = _recording_config_key(name) if recording_config else ""

    s3 = _s3()
    s3.put_object(Bucket=bkt, Key=py_key, Body=py_bytes, ContentType="text/x-python")
    s3.put_object(Bucket=bkt, Key=params_key, Body=params_bytes, ContentType=ct)
    if recording_config:
        s3.put_object(
            Bucket=bkt,
            Key=recording_config_key,
            Body=json.dumps(recording_config, indent=2, sort_keys=True).encode("utf-8"),
            ContentType="application/json",
        )

    db_error = ""
    db_result: dict[str, Any] = {}
    try:
        db_result = upsert_recorded_flow(
            name=name, file_name=py_key, data_file_name=params_key,
            start_url=start_url, user_id=(body.user_id or DEFAULT_USER_ID), overwrite=body.overwrite,
        )
    except Exception as exc:  # MinIO already succeeded; surface DB issue without failing upload
        db_error = f"{type(exc).__name__}: {exc}"

    return {
        "ok": True, "bucket": bkt, "name": name, "py_key": py_key, "params_key": params_key,
        "start_url": start_url, "param_rows": len(param_sets), "missing_placeholders": missing,
        "multi_line_row_count": len(multi_line_rows),
        "recording_config_key": recording_config_key, "recording_config": recording_config,
        "db": db_result, "db_error": db_error,
        "run_cmd": _run_cmd(name, py_key, after_action_wait_ms=DEFAULT_AFTER_ACTION_WAIT_MS),
    }


def _run_payload(
    name: str,
    py_key: str,
    *,
    execution_mode: str = "parallel",
    after_action_wait_ms: int = DEFAULT_AFTER_ACTION_WAIT_MS,
) -> dict[str, Any]:
    payload = {
        "test_suite_id": name,
        "recordings": [
            {
                "id": name,
                "name": name,
                "file": py_key,
                "after_action_wait_ms": after_action_wait_ms,
            }
        ],
        "execution_mode": execution_mode,
    }
    return payload


def _run_cmd(
    name: str,
    py_key: str,
    execution_mode: str = "parallel",
    *,
    after_action_wait_ms: int = DEFAULT_AFTER_ACTION_WAIT_MS,
) -> str:
    payload = _run_payload(
        name,
        py_key,
        execution_mode=execution_mode,
        after_action_wait_ms=after_action_wait_ms,
    )
    return f"{AETHERION_BIN} agent 'ACT Agent' '{json.dumps(payload, separators=(',', ':'))}'"


class RunBody(BaseModel):
    name: str
    parameters: Any = {}
    execution_mode: str = "parallel"
    after_action_wait_ms: int | None = DEFAULT_AFTER_ACTION_WAIT_MS
    task_queue: str = ""
    download_report: bool = True


class SuiteRecording(BaseModel):
    name: str
    parameters: Any = None  # optional per-recording inline params; else saved workbook is used


class SuiteRunBody(BaseModel):
    recordings: list[SuiteRecording]
    suite_id: str = ""
    # Suites default to sequential so flow-context chaining (one recording's
    # extract/ai_extract feeding the next) actually works.
    execution_mode: str = "sequential"
    after_action_wait_ms: int | None = DEFAULT_AFTER_ACTION_WAIT_MS
    task_queue: str = ""
    download_report: bool = True


def _extract_agent_result(stdout: str) -> Any:
    for chunk in re.findall(r"\{.*\}|\[.*\]", stdout, re.DOTALL):
        try:
            return json.loads(chunk)
        except Exception:
            continue
    return None


def _find_report_key(result: Any) -> str:
    items = result if isinstance(result, list) else (result.get("result") or result.get("outputs") or [] if isinstance(result, dict) else [])
    if isinstance(result, dict) and not items:
        items = [result]
    for item in items if isinstance(items, list) else []:
        if isinstance(item, dict) and str(item.get("type") or "") == "s3_download_link":
            if item.get("file_key"):
                return str(item["file_key"])
    return ""


def _report_url_for_dest(dest: Path) -> str:
    return f"/downloads/{dest.name}"


def _build_recording_entry(
    name: str,
    *,
    parameters: Any = None,
    after_action_wait_ms: int,
    bucket: str,
) -> dict[str, Any]:
    """Build one `recordings[]` entry for the agent payload.

    Inline `parameters` win; otherwise the recording's saved workbook is loaded
    and shipped explicitly so the worker does not re-parse it remotely. Shared by
    the single-run and suite-run endpoints so both resolve params identically.
    """
    safe = _safe_name(name)
    py_key = f"recordings/{safe}/{safe}.py"
    rec: dict[str, Any] = {"id": safe, "name": safe, "file": py_key}
    if parameters:
        inline_multi_line_rows = _normalize_multi_line_rows(parameters)
        inline_parameter_payload = _payload_without_multi_line(parameters)
        try:
            rec["parameters"] = normalize_param_sets(inline_parameter_payload)[0]
        except ValueError as exc:
            raise HTTPException(400, f"{safe}: {exc}")
        if inline_multi_line_rows:
            rec[DEFAULT_MULTI_LINE_SHEET_NAME] = inline_multi_line_rows
        rec["skip_parameters_file_load"] = True
    else:
        params_rows, multi_line_rows, _params_key = _load_saved_runtime_payload(safe, bucket)
        if params_rows or multi_line_rows:
            rec["parameters"] = dict(params_rows[0]) if params_rows else {}
            if multi_line_rows:
                rec[DEFAULT_MULTI_LINE_SHEET_NAME] = list(multi_line_rows)
            rec["skip_parameters_file_load"] = True
    rec["after_action_wait_ms"] = after_action_wait_ms
    return rec


def _build_recording_entries(
    name: str,
    *,
    parameters: Any = None,
    after_action_wait_ms: int,
    bucket: str,
) -> list[dict[str, Any]]:
    safe = _safe_name(name)
    py_key = f"recordings/{safe}/{safe}.py"

    if parameters:
        inline_multi_line_rows = _normalize_multi_line_rows(parameters)
        inline_parameter_payload = _payload_without_multi_line(parameters)
        try:
            params_rows = normalize_param_sets(inline_parameter_payload, allow_empty=True)
        except ValueError as exc:
            raise HTTPException(400, f"{safe}: {exc}")
        multi_line_rows = inline_multi_line_rows
        params_key = ""
    else:
        params_rows, multi_line_rows, params_key = _load_saved_runtime_payload(safe, bucket)

    if not params_rows:
        params_rows = [{}]

    prepared_rows: list[dict[str, Any]] = []
    if len(params_rows) == 1:
        prepared_rows.append(
            {
                "parameters": dict(params_rows[0]),
                DEFAULT_MULTI_LINE_SHEET_NAME: list(multi_line_rows),
                "parameter_row_index": 1,
                "parameter_set_index": 1,
            }
        )
    elif not multi_line_rows:
        for index, header_row in enumerate(params_rows, start=1):
            prepared_rows.append(
                {
                    "parameters": dict(header_row),
                    DEFAULT_MULTI_LINE_SHEET_NAME: [],
                    "parameter_row_index": index,
                    "parameter_set_index": index,
                }
            )
    else:
        match_key = MATCH_KEY

        grouped_lines: dict[str, list[dict[str, str]]] = {}
        for line_index, line_row in enumerate(multi_line_rows, start=1):
            line_match_value = _extract_row_value_by_key(line_row, match_key)
            if not line_match_value:
                raise HTTPException(
                    400,
                    f"{safe}: multi_line row {line_index} is missing match key '{match_key}'.",
                )
            grouped_lines.setdefault(line_match_value, []).append(dict(line_row))

        seen_header_values: set[str] = set()
        for index, header_row in enumerate(params_rows, start=1):
            header_match_value = _extract_row_value_by_key(header_row, match_key)
            if not header_match_value:
                raise HTTPException(
                    400,
                    f"{safe}: params row {index} is missing match key '{match_key}'.",
                )
            if header_match_value in seen_header_values:
                raise HTTPException(
                    400,
                    f"{safe}: duplicate params match key '{header_match_value}' for '{match_key}'.",
                )
            seen_header_values.add(header_match_value)
            matched_lines = grouped_lines.pop(header_match_value, [])
            if not matched_lines:
                raise HTTPException(
                    400,
                    f"{safe}: no multi_line rows found for params row {index} using '{match_key}={header_match_value}'.",
                )
            prepared_rows.append(
                {
                    "parameters": dict(header_row),
                    DEFAULT_MULTI_LINE_SHEET_NAME: matched_lines,
                    "parameter_row_index": index,
                    "parameter_set_index": index,
                }
            )

        if grouped_lines:
            extra_keys = ", ".join(sorted(grouped_lines))
            raise HTTPException(
                400,
                f"{safe}: multi_line rows contain unmatched '{match_key}' values: {extra_keys}.",
            )

    entries: list[dict[str, Any]] = []
    total_rows = len(prepared_rows)
    for prepared in prepared_rows:
        row_index = int(prepared.get("parameter_row_index") or 1)
        entry_name = safe if total_rows == 1 else f"{safe} [row {row_index}]"
        entry_id = safe if total_rows == 1 else f"{safe}-row-{row_index}"
        rec: dict[str, Any] = {
            "id": entry_id,
            "name": entry_name,
            "file": py_key,
            "after_action_wait_ms": after_action_wait_ms,
            "parameter_row_index": row_index,
            "parameter_set_index": int(prepared.get("parameter_set_index") or row_index),
        }
        parameters_row = prepared.get("parameters")
        multi_line_row_values = prepared.get(DEFAULT_MULTI_LINE_SHEET_NAME)
        if isinstance(parameters_row, dict):
            rec["parameters"] = dict(parameters_row)
            rec["skip_parameters_file_load"] = True
        if isinstance(multi_line_row_values, list) and multi_line_row_values:
            rec[DEFAULT_MULTI_LINE_SHEET_NAME] = list(multi_line_row_values)
            rec["skip_parameters_file_load"] = True
        if params_key:
            rec["parameters_file_key"] = params_key
        entries.append(rec)

    return entries


def _submit_agent_payload(
    payload: dict[str, Any],
    *,
    task_queue: str = "",
) -> tuple[list[str], "subprocess.CompletedProcess[str]"]:
    """Shell out to the local aetherion CLI and wait for the run to finish."""
    cmd = [AETHERION_BIN, "agent", "ACT Agent", json.dumps(payload), "--wait"]
    if task_queue:
        cmd += ["--task-queue", task_queue]
    if not Path(AETHERION_BIN).exists():
        raise HTTPException(400, f"aetherion CLI not found at {AETHERION_BIN}")
    env = _build_local_aetherion_cli_env()
    try:
        proc = subprocess.run(
            cmd,
            cwd=str(TEST_RUNNER_DIR),
            env=env,
            capture_output=True,
            text=True,
            timeout=2400,
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(504, "run timed out after 40 min")
    return cmd, proc


def _download_run_report(report_key: str, label: str) -> tuple[str, str]:
    """Download the run's HTML report to a unique local file; return (path, url).

    The S3 key is always <suite>/<run_id>/report.html, so naming the local file by
    basename alone makes every run collide on downloads/report.html at the same URL
    — the browser then serves a cached previous run. Key it by run_id so each run
    gets a unique URL and "Open report" always shows the run just executed.
    """
    try:
        DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
        raw = _s3().get_object(Bucket=BUCKET, Key=report_key)["Body"].read()
        run_uuid = Path(report_key).parent.name or "run"
        dest = DOWNLOADS_DIR / f"{_safe_name(label)}_{run_uuid}.html"
        dest.write_bytes(raw)
        return str(dest), _report_url_for_dest(dest)
    except Exception as exc:
        return f"(download failed: {exc})", ""


@app.post("/api/run")
def run(body: RunBody):
    name = _safe_name(body.name)
    effective_wait_ms = DEFAULT_AFTER_ACTION_WAIT_MS if body.after_action_wait_ms is None else int(body.after_action_wait_ms)
    entries = _build_recording_entries(
        name,
        parameters=body.parameters if body.parameters else None,
        after_action_wait_ms=effective_wait_ms,
        bucket=BUCKET,
    )
    payload = {"test_suite_id": name, "recordings": entries, "execution_mode": body.execution_mode}
    cmd, proc = _submit_agent_payload(payload, task_queue=body.task_queue)
    stdout, stderr = proc.stdout, proc.stderr
    agent_result = _extract_agent_result(stdout) or _extract_agent_result(stderr)
    report_key = _find_report_key(agent_result)
    report_local, report_url = "", ""
    if report_key and body.download_report:
        report_local, report_url = _download_run_report(report_key, name)
    inline_parameters = entries[0].get("parameters") if entries and isinstance(entries[0].get("parameters"), dict) else {}
    inline_multi_line_rows = (
        entries[0].get(DEFAULT_MULTI_LINE_SHEET_NAME)
        if entries and isinstance(entries[0].get(DEFAULT_MULTI_LINE_SHEET_NAME), list)
        else []
    )
    return {
        "ok": proc.returncode == 0,
        "returncode": proc.returncode,
        "cmd": " ".join(cmd[:3]) + " '<payload>' --wait",
        "stdout": stdout[-20000:],
        "stderr": stderr[-8000:],
        "report_key": report_key,
        "report_local": report_local,
        "report_url": report_url,
        "used_after_action_wait_ms": effective_wait_ms,
        "execution_mode": body.execution_mode,
        "inline_parameter_keys": sorted(str(key) for key in inline_parameters),
        "inline_multi_line_row_count": len(inline_multi_line_rows),
        "prepared_recording_count": len(entries),
        "prepared_recording_names": [str(entry.get("name") or "") for entry in entries],
    }


@app.post("/api/run-suite")
def run_suite(body: SuiteRunBody):
    if not body.recordings:
        raise HTTPException(400, "at least one recording is required for a suite")
    effective_wait_ms = DEFAULT_AFTER_ACTION_WAIT_MS if body.after_action_wait_ms is None else int(body.after_action_wait_ms)
    entries = [
        entry
        for r in body.recordings
        for entry in _build_recording_entries(
            r.name,
            parameters=r.parameters if r.parameters else None,
            after_action_wait_ms=effective_wait_ms,
            bucket=BUCKET,
        )
    ]
    names = [e["name"] for e in entries]
    suite_id = _safe_name(body.suite_id) or _safe_name("suite_" + "_".join(names[:2]))[:80] or "suite"
    payload = {"test_suite_id": suite_id, "recordings": entries, "execution_mode": body.execution_mode}
    cmd, proc = _submit_agent_payload(payload, task_queue=body.task_queue)
    stdout, stderr = proc.stdout, proc.stderr
    agent_result = _extract_agent_result(stdout) or _extract_agent_result(stderr)
    report_key = _find_report_key(agent_result)
    report_local, report_url = "", ""
    if report_key and body.download_report:
        report_local, report_url = _download_run_report(report_key, suite_id)
    return {
        "ok": proc.returncode == 0,
        "returncode": proc.returncode,
        "cmd": " ".join(cmd[:3]) + " '<payload>' --wait",
        "stdout": stdout[-20000:],
        "stderr": stderr[-8000:],
        "report_key": report_key,
        "report_local": report_local,
        "report_url": report_url,
        "used_after_action_wait_ms": effective_wait_ms,
        "execution_mode": body.execution_mode,
        "suite_id": suite_id,
        "recordings": names,
    }


@app.get("/api/config")
def get_config():
    return {"bucket": BUCKET, "storage_endpoint": STORAGE_ENDPOINT, "aetherion_bin": AETHERION_BIN,
            "test_runner_dir": str(TEST_RUNNER_DIR), "pg": {"host": PG["host"], "port": PG["port"], "db": PG["dbname"]},
            "default_after_action_wait_ms": DEFAULT_AFTER_ACTION_WAIT_MS}


@app.get("/", response_class=HTMLResponse)
def index():
    return HTMLResponse(HTML, headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0", "Pragma": "no-cache", "Expires": "0"})


def _find_available_port(preferred_port: int, host: str = "127.0.0.1", *, attempts: int = 20) -> int:
    for port in range(preferred_port, preferred_port + max(1, attempts)):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                sock.bind((host, port))
            except OSError:
                continue
            return port
    raise RuntimeError(
        f"Could not find a free port starting at {preferred_port} on {host} after {attempts} attempts."
    )


HTML = """<!doctype html><html><head><meta charset=utf-8><meta http-equiv="Cache-Control" content="no-store, no-cache, must-revalidate, max-age=0"><meta http-equiv="Pragma" content="no-cache"><meta http-equiv="Expires" content="0"><title>agent_shubham</title>
<style>
:root{--bg:#0f1117;--bg-soft:#12141b;--panel:#171a22;--panel-2:#141821;--line:#262b36;--line-strong:#39507d;--fg:#e6e9ef;--mut:#8b93a7;--acc:#4f8cff;--acc-2:#6b98ff;--ok:#36c98a;--bad:#ff6b6b;--shadow:0 14px 30px rgba(0,0,0,.22)}
*{box-sizing:border-box}
html{scroll-behavior:smooth}
body{margin:0;font:14px/1.5 Inter,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--bg);color:var(--fg)}
header{display:flex;align-items:center;gap:12px;justify-content:space-between;padding:12px 18px;border-bottom:1px solid var(--line);background:var(--panel);position:sticky;top:0;z-index:5}
header h1{font-size:15px;margin:0;font-weight:700;letter-spacing:.01em}
header .meta{color:var(--mut);font-size:12px}
.wrap{display:grid;grid-template-columns:380px 1fr;height:calc(100vh - 56px)}
.col{overflow:auto;padding:14px}
.col.left{border-right:1px solid var(--line);background:var(--bg-soft)}
.card{background:var(--panel);border:1px solid var(--line);border-radius:14px;padding:14px;margin-bottom:14px;box-shadow:var(--shadow);animation:fadeUp .24s cubic-bezier(.22,1,.36,1);transition:transform .18s ease,border-color .18s ease,box-shadow .18s ease}
.card:hover{transform:translateY(-1px);border-color:var(--line-strong);box-shadow:0 22px 44px rgba(5,10,20,.32)}
.script{display:grid;grid-template-columns:minmax(0,1fr) 92px;gap:8px;align-items:center;padding:9px 10px;border:1px solid var(--line);border-radius:10px;margin-bottom:6px;background:var(--bg-soft);transition:transform .16s ease,border-color .16s ease,background .16s ease,box-shadow .16s ease;animation:fadeUp .18s ease}
.script:hover{border-color:var(--line-strong);transform:translateY(-1px);box-shadow:0 10px 24px rgba(4,8,18,.28)}
.script.sel{border-color:var(--acc);background:#13203a;box-shadow:0 0 0 1px rgba(79,140,255,.18),0 14px 30px rgba(7,12,23,.24)}
.script-main{min-width:0}
.script-top{display:flex;gap:8px;align-items:flex-start;min-width:0}
.script-copy{min-width:0}
.script .nm{font-weight:700;font-size:13px;word-break:normal;overflow-wrap:anywhere;line-height:1.45}
.script .sub{color:var(--mut);font-size:11px;line-height:1.4;margin-top:2px}
.script-actions{display:grid;grid-template-columns:1fr;gap:6px;align-items:stretch;width:92px}
.script-btn{width:100%;min-width:0;height:34px;padding:0 10px;display:inline-flex;align-items:center;justify-content:center;border-radius:9px}
.badge{display:inline-flex;align-items:center;gap:4px;font-size:10px;padding:2px 7px;border-radius:999px;border:1px solid var(--line-strong);background:#0f1628}
.badge.db{color:var(--ok);border-color:rgba(56,211,159,.28);background:rgba(56,211,159,.08)}
.badge.nodb{color:var(--mut)}
button{background:var(--acc);color:#fff;border:0;border-radius:10px;padding:8px 12px;font-weight:700;cursor:pointer;font-size:12px;line-height:1;transition:transform .14s ease,filter .14s ease,box-shadow .14s ease;box-shadow:0 8px 18px rgba(32,76,176,.20)}
button:hover{transform:translateY(-1px);filter:brightness(1.03)}
button:active{transform:translateY(0)}
button.ghost{background:#19233a;color:var(--fg);box-shadow:none;border:1px solid var(--line)}
button.ghost:hover{border-color:var(--line-strong);background:#1d2943}
button:disabled{opacity:.55;cursor:wait;transform:none;filter:none}
a.btn{display:inline-flex;align-items:center;justify-content:center;background:var(--acc);color:#fff;border:0;border-radius:10px;padding:8px 12px;font-weight:700;cursor:pointer;font-size:12px;text-decoration:none;line-height:1;box-shadow:0 8px 18px rgba(32,76,176,.20);transition:transform .14s ease,filter .14s ease}
a.btn:hover{transform:translateY(-1px);filter:brightness(1.03)}
a.btn.ghost{background:#19233a;color:var(--fg);box-shadow:none;border:1px solid var(--line)}
label{display:block;color:var(--mut);font-size:12px;margin:10px 0 4px}
input,textarea,select{width:100%;background:#0b1222;color:var(--fg);border:1px solid var(--line);border-radius:10px;padding:9px 10px;font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;font-size:12px;transition:border-color .14s ease,box-shadow .14s ease,background .14s ease}
input:focus,textarea:focus,select:focus{outline:none;border-color:var(--acc);box-shadow:0 0 0 3px rgba(91,140,255,.16);background:#0e1629}
textarea{resize:vertical}
.row{display:flex;gap:12px}
.row>*{flex:1}.form-top-row{align-items:flex-start}.top-field{display:flex;flex-direction:column;min-width:0}.top-field label{margin-top:0}.top-field input,.top-field select{height:38px;padding-top:0;padding-bottom:0}.top-actions-field{flex:0 0 auto;min-width:max-content}.top-actions-field label{visibility:hidden}.top-actions-field .form-top-actions{display:flex;align-items:center;justify-content:flex-end;gap:8px}.top-actions-field .form-top-actions .ghost,.top-actions-field .form-top-actions button{height:38px;min-width:auto;padding:0 14px;box-shadow:none}.prompt-strip{margin:2px 0 10px}.prompt-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:6px}.prompt-head label{margin:0}.prompt-head .inline-check.compact-check{flex:0 0 auto;margin:0;padding:0;border:0;background:transparent;border-radius:0;justify-content:flex-end;min-height:auto}.prompt-head .inline-check.compact-check input{margin:0}.repeatable-note{margin-top:6px;line-height:1.45}.form-top-row{align-items:flex-start}.top-field{display:flex;flex-direction:column;min-width:0}.top-field label{margin-top:0}.top-field input,.top-field select{height:38px;padding-top:0;padding-bottom:0}.top-actions-field{flex:0 0 auto;min-width:max-content}.top-actions-field label{visibility:hidden}.top-actions-field .form-top-actions{display:flex;align-items:center;justify-content:flex-end;gap:8px}.top-actions-field .form-top-actions .ghost,.top-actions-field .form-top-actions button{height:38px;min-width:auto;padding:0 14px;box-shadow:none}.prompt-strip{margin:2px 0 10px}.prompt-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:6px}.prompt-head label{margin:0}.prompt-head .inline-check.compact-check{flex:0 0 auto;margin:0;padding:0;border:0;background:transparent;border-radius:0;justify-content:flex-end;min-height:auto}.prompt-head .inline-check.compact-check input{margin:0}.repeatable-note{margin-top:6px;line-height:1.45}.form-top-row{align-items:flex-start}.top-field{display:flex;flex-direction:column;min-width:0}.top-field label{margin-top:0}.top-field input,.top-field select{height:38px;padding-top:0;padding-bottom:0}.top-actions-field{flex:0 0 auto;min-width:max-content}.top-actions-field label{visibility:hidden}.top-actions-field .form-top-actions{display:flex;align-items:center;justify-content:flex-end;gap:8px}.top-actions-field .form-top-actions .ghost,.top-actions-field .form-top-actions button{height:38px;min-width:auto;padding:0 14px;box-shadow:none}.prompt-strip{margin:2px 0 10px}.prompt-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:6px}.prompt-head label{margin:0}.prompt-head .inline-check.compact-check{flex:0 0 auto;margin:0;padding:0;border:0;background:transparent;border-radius:0;justify-content:flex-end;min-height:auto}.prompt-head .inline-check.compact-check input{margin:0}.repeatable-note{margin-top:6px;line-height:1.45}.form-top-row{align-items:flex-start}.top-field{display:flex;flex-direction:column;min-width:0}.top-field label{margin-top:0}.top-field input,.top-field select{height:38px;padding-top:0;padding-bottom:0}.top-actions-field{flex:0 0 auto;min-width:max-content}.top-actions-field label{visibility:hidden}.top-actions-field .form-top-actions{display:flex;align-items:center;justify-content:flex-end;gap:8px}.top-actions-field .form-top-actions .ghost,.top-actions-field .form-top-actions button{height:38px;min-width:auto;padding:0 14px;box-shadow:none}.prompt-strip{margin:2px 0 10px}.prompt-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:6px}.prompt-head label{margin:0}.prompt-head .inline-check.compact-check{flex:0 0 auto;margin:0;padding:0;border:0;background:transparent;border-radius:0;justify-content:flex-end;min-height:auto}.prompt-head .inline-check.compact-check input{margin:0}.repeatable-note{margin-top:6px;line-height:1.45}.form-top-row{align-items:flex-start}.top-field{display:flex;flex-direction:column;min-width:0}.top-field label{margin-top:0}.top-field input,.top-field select{height:38px;padding-top:0;padding-bottom:0}.top-actions-field{flex:0 0 auto;min-width:max-content}.top-actions-field label{visibility:hidden}.form-top-actions{display:flex;align-items:center;justify-content:flex-end;gap:8px}.form-top-actions .ghost,.form-top-actions button{height:38px;min-width:auto;padding:0 14px;box-shadow:none}.prompt-strip{margin:2px 0 10px}.prompt-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:6px}.prompt-head label{margin:0}.prompt-head .compact-check{flex:0 0 auto}.compact-check{margin:0;padding:0;border:0;background:transparent;justify-content:flex-end;min-height:auto}.compact-check input{margin:0}.repeatable-note{margin-top:6px;line-height:1.45}.form-top-row{align-items:flex-start}.top-field{display:flex;flex-direction:column;min-width:0}.top-field label{margin-top:0}.top-field input,.top-field select{height:38px;padding-top:0;padding-bottom:0}.top-actions-field{flex:0 0 auto;min-width:max-content}.top-actions-field label{visibility:hidden}.form-top-actions{display:flex;align-items:center;justify-content:flex-end;gap:8px}.form-top-actions .ghost,.form-top-actions button{height:38px;min-width:auto;padding:0 14px;box-shadow:none}.prompt-strip{margin:2px 0 10px}.prompt-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:6px}.prompt-head label{margin:0}.prompt-head .compact-check{flex:0 0 auto}.compact-check{margin:0;padding:0;border:0;background:transparent;justify-content:flex-end;min-height:auto}.compact-check input{margin:0}.repeatable-note{margin-top:6px;line-height:1.45}
.inline-check{display:flex;align-items:center;gap:10px;margin:14px 0 6px;color:var(--fg);font-size:12px;padding:10px 12px;border:1px solid var(--line);border-radius:12px;background:#10182a}
.inline-check input{width:auto;margin:0}
.tabs{display:flex;gap:8px;margin-bottom:12px}
.tab{padding:8px 13px;border:1px solid var(--line);border-radius:10px;cursor:pointer;color:var(--mut);background:#10182a;transition:border-color .14s ease,background .14s ease,color .14s ease}
.tab.on{color:var(--fg);border-color:var(--acc);background:#16213a}
pre{background:#0a101d;border:1px solid var(--line);border-radius:10px;padding:12px;overflow:auto;font-size:12px;white-space:pre-wrap;max-height:50vh}
.msg{padding:10px 12px;border-radius:10px;margin:10px 0;font-size:12px;display:none}
.msg.ok{background:#11251d;border:1px solid rgba(56,211,159,.26);color:var(--ok);display:block}
.msg.err{background:#291619;border:1px solid rgba(255,113,113,.28);color:var(--bad);display:block}
a{color:var(--acc-2)}
.spin{display:inline-block;width:12px;height:12px;border:2px solid #fff5;border-top-color:#fff;border-radius:50%;animation:s .7s linear infinite;vertical-align:-2px}
.actions{display:flex;gap:8px;flex-wrap:wrap;margin:12px 0 0}
.meta-line{color:var(--mut);font-size:12px;margin-top:8px;line-height:1.5}
@keyframes s{to{transform:rotate(360deg)}}
@keyframes fadeUp{from{opacity:0;transform:translateY(4px)}to{opacity:1;transform:translateY(0)}}
.suitechk,.inline-check input[type=checkbox]{-webkit-appearance:none;appearance:none;width:18px;height:18px;border:1px solid var(--line-strong);background:#0d1425;border-radius:6px;cursor:pointer;flex:0 0 auto;margin:0;padding:0;position:relative;transition:border-color .14s ease,background .14s ease,box-shadow .14s ease}
.suitechk::after,.inline-check input[type=checkbox]::after{content:"";position:absolute;left:5px;top:2px;width:4px;height:8px;border:2px solid transparent;border-top:0;border-left:0;transform:rotate(45deg);opacity:0}
.suitechk:checked,.inline-check input[type=checkbox]:checked{background:var(--acc);border-color:var(--acc);box-shadow:0 0 0 3px rgba(91,140,255,.14)}
.suitechk:checked::after,.inline-check input[type=checkbox]:checked::after{border-color:#fff;opacity:1}
.suitechk:hover,.inline-check input[type=checkbox]:hover{border-color:var(--acc)}
.CodeMirror{border:1px solid var(--line);border-radius:12px;height:auto;font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;font-size:12.5px;line-height:1.55;box-shadow:inset 0 1px 0 rgba(255,255,255,.02)}
.CodeMirror-focused{border-color:var(--acc);box-shadow:0 0 0 3px rgba(91,140,255,.16)}
.CodeMirror-gutters{border-right:1px solid var(--line)}
.editor-shell{margin:6px 0 10px;border:1px solid var(--line);border-radius:14px;overflow:hidden;background:#0b1222;box-shadow:inset 0 1px 0 rgba(255,255,255,.02);transition:border-color .14s ease,box-shadow .14s ease,transform .14s ease}
.editor-shell:hover{border-color:var(--line-strong)}
.editor-shell:focus-within{border-color:var(--acc);box-shadow:0 0 0 3px rgba(91,140,255,.16)}
.editor-shell.valid{border-color:rgba(56,211,159,.34)}
.editor-shell.invalid{border-color:rgba(255,113,113,.38)}
.editor-head{display:flex;align-items:center;justify-content:space-between;gap:12px;min-height:38px;padding:6px 12px;border-bottom:1px solid var(--line);background:var(--panel-2)}.editor-json-status{font-size:11px;max-width:260px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--mut)}
.editor-head-left{display:flex;align-items:center;gap:10px;min-width:0}.editor-head-right{display:flex;align-items:center;gap:8px;min-width:0}.editor-head-right{display:flex;align-items:center;gap:8px;min-width:0}.editor-head-right{display:flex;align-items:center;gap:8px;min-width:0}.editor-head-right{display:flex;align-items:center;gap:8px;min-width:0}.editor-head-right{display:flex;align-items:center;gap:8px;min-width:0}.editor-head-right{display:flex;align-items:center;gap:8px;min-width:0}.editor-head-right{display:flex;align-items:center;gap:8px;min-width:0}.editor-head-right{display:flex;align-items:center;gap:8px;min-width:0}.editor-icon-actions{display:flex;align-items:center;gap:6px}.editor-icon-btn{width:28px;height:28px;padding:0;display:inline-flex;align-items:center;justify-content:center;border-radius:8px;font-size:13px;box-shadow:none}
.editor-kind{font-size:11px;font-weight:700;color:var(--fg);letter-spacing:.02em;text-transform:uppercase}
.editor-hint{font-size:11px;color:var(--mut);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.editor-status{font-size:11px;color:var(--mut);white-space:nowrap}
.editor-pane{display:grid;grid-template-columns:48px minmax(0,1fr);align-items:stretch;min-height:0}
.editor-gutter{padding:12px 8px 12px 0;background:#0d1425;border-right:1px solid var(--line);color:#5f7097;font:12px/1.55 ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;text-align:right;user-select:none;overflow:hidden}
.editor-gutter span{display:block}
.editor-shell textarea.code-fallback{display:block;width:100%;margin:0;border:0;border-radius:0;padding:12px 14px;background:#0b1222;color:var(--fg);font:12.5px/1.55 ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;white-space:pre;overflow:auto;resize:vertical;tab-size:2;outline:none;box-shadow:none}
.editor-shell textarea.code-fallback:focus{background:#0b1222;box-shadow:none}
.editor-shell.cm-enhanced .editor-pane{display:block}
.editor-shell.cm-enhanced .editor-gutter{display:none}
.editor-shell.cm-enhanced .CodeMirror{border:0;border-radius:0;box-shadow:none}
.editor-shell.cm-enhanced .CodeMirror-focused{border-color:transparent;box-shadow:none}
.editor-grid{display:grid;grid-template-columns:minmax(0,1.1fr) minmax(0,.95fr);gap:14px;align-items:start;margin:10px 0 8px}
.field-block{min-width:0;display:flex;flex-direction:column}
.field-head{display:flex;align-items:center;justify-content:space-between;gap:12px;height:32px;margin-bottom:6px}
.field-head label{margin:0;line-height:1.2;flex:1 1 auto}
.field-actions{display:flex;align-items:center;gap:8px;flex-wrap:wrap;justify-content:flex-end}
.field-actions .ghost{box-shadow:none}.editor-top-actions{display:flex;align-items:flex-end;justify-content:flex-end;gap:8px;flex:0 0 auto}.editor-top-actions .ghost{height:36px}.editor-top-actions{display:flex;align-items:flex-end;justify-content:flex-end;gap:8px;flex:0 0 auto}.editor-top-actions .ghost{height:36px}.editor-top-actions{display:flex;align-items:flex-end;justify-content:flex-end;gap:8px;flex:0 0 auto}.editor-top-actions .ghost{height:36px}.editor-top-actions{display:flex;align-items:flex-end;justify-content:flex-end;gap:8px;flex:0 0 auto}.editor-top-actions .ghost{height:36px}.editor-top-actions{display:flex;align-items:flex-end;justify-content:flex-end;gap:8px;flex:0 0 auto}.editor-top-actions .ghost{height:36px}.editor-top-actions{display:flex;align-items:flex-end;justify-content:flex-end;gap:8px;flex:0 0 auto}.editor-top-actions .ghost{height:36px}.editor-top-actions{display:flex;align-items:flex-end;justify-content:flex-end;gap:8px;flex:0 0 auto}.editor-top-actions .ghost{height:36px}.form-top-actions{display:flex;align-items:flex-end;justify-content:flex-end;gap:8px;flex:0 0 auto}.form-top-actions .ghost,.form-top-actions button{height:36px}.field-actions-placeholder{visibility:hidden;pointer-events:none}
.editor-tools{display:flex;align-items:center;justify-content:flex-end;gap:8px;margin:8px 0 4px}
.editor-tools .pill{font-size:11px;color:var(--mut)}
.editor-compact-btn{padding:6px 10px;min-width:auto;box-shadow:none}
.suite-card{padding:12px}
.suite-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}
.suite-list{margin:0 0 10px 0;padding:0;list-style:none;display:grid;gap:7px}
.suite-item{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:8px 10px;border:1px solid var(--line);border-radius:10px;background:#0f1628}
.suite-name{word-break:break-word;font-size:12px}
.suite-actions{display:flex;gap:4px;white-space:nowrap}
.suite-actions button{padding:6px 8px;min-width:auto;box-shadow:none}
.section-row{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}
.search{margin-bottom:8px}
.refresh-btn{padding:7px 10px;min-width:auto}.view-switch{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px;margin-bottom:8px}.side-tab{display:flex;align-items:center;justify-content:center;gap:8px;height:36px;padding:0 10px;box-shadow:none}.side-tab.on{background:#1e2b46;border:1px solid var(--acc);color:var(--fg)}.view-count{display:inline-flex;align-items:center;justify-content:center;min-width:20px;height:20px;padding:0 6px;border-radius:999px;background:#0f1628;border:1px solid var(--line);font-size:11px;color:var(--mut)}.side-tab.on .view-count{border-color:rgba(79,140,255,.45);color:var(--fg)}.run-shell{padding:12px}.run-tabbar{display:flex;gap:8px;overflow:auto;padding-bottom:2px;margin:10px 0}.run-tab{min-width:0;max-width:220px;height:34px;padding:0 12px;border-radius:9px;background:#11192b;color:var(--mut);border:1px solid var(--line);box-shadow:none;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.run-tab.on{background:#1d2a44;border-color:var(--acc);color:var(--fg)}.run-empty{padding:14px 12px;border:1px dashed var(--line);border-radius:10px;color:var(--mut);background:#111827}.run-title{display:flex;justify-content:space-between;align-items:center;margin:2px 0 8px 0}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;margin-top:14px}
.upload-actions button{min-height:40px}
@media(max-width:860px){.editor-grid{grid-template-columns:1fr}.field-actions{justify-content:flex-start}.prompt-head{flex-direction:column;align-items:flex-start}.prompt-head .inline-check.compact-check{justify-content:flex-start}}
@media(max-width:980px){.wrap{grid-template-columns:1fr}.col.left{border-right:0;border-bottom:1px solid var(--line)}.row{flex-direction:column;gap:10px}.script{grid-template-columns:1fr}.script-actions{grid-template-columns:repeat(2,minmax(0,1fr));width:100%}}
</style>
<link rel=stylesheet href="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/codemirror.min.css">
<link rel=stylesheet href="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/theme/dracula.min.css">
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/codemirror.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/mode/python/python.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/mode/javascript/javascript.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/addon/edit/matchbrackets.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/addon/edit/closebrackets.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/addon/selection/active-line.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.16/addon/search/searchcursor.min.js"></script>
</head><body>
<header><h1>agent_shubham</h1><span class=meta id=cfg></span><span style=flex:1></span><button class="ghost refresh-btn" onclick=loadScripts()>Refresh</button></header>
<div class=wrap>
  <div class="col left">
    <div class=section-row><b>Recordings</b><span class=meta id=count></span></div>
    <div class=view-switch><button class="ghost side-tab on" id=side-edit onclick="tab('edit')">Upload / Edit</button><button class="ghost side-tab" id=side-run onclick="tab('run')">Run Output <span id=runTabCount class=view-count>0</span></button></div>
    <input id=scriptSearch class=search placeholder="search script name">
    <div id=suiteTray style=display:none></div>
    <div id=list></div>
  </div>
  <div class="col main-panel">
    <div id=p-edit>
      <div class=card>
        <div class="row form-top-row"><div class=top-field><label>Recording name</label><input id=name placeholder="recording_name_v1"></div>
        <div class=top-field style=flex:0.4><label>Overwrite</label><select id=ov><option value=true>true</option><option value=false>false</option></select></div>
        <div class="top-field top-actions-field"><label>&nbsp;</label><div class=form-top-actions><button id=upbtn onclick=doUpload()>Upload to MinIO + DB</button><button class=ghost onclick=clearForm()>Clear</button></div></div></div>
        <div class=editor-grid>
          <div class=field-block>
            <textarea id=py rows=14 placeholder="paste a recorded Playwright or plain Python script"></textarea>
          </div>
          <div class=field-block>
            <textarea id=params rows=14 placeholder='{"params":[{"username":"...","password":"..."}]}'></textarea>
          </div>
        </div>
        <label>Prompt</label><textarea id=prompt rows=4 placeholder="Recording guidance. If it has a repeatable block, name the repeated fields here, e.g. 'Repeat the line item for each row: Description, Quantity, Unit Price'."></textarea>
        <label class=inline-check><input id=repeatableLineItemsEnabled type=checkbox onchange=toggleRepeatableLineItems()>This recording has a repeatable block</label>
        <div id=repeatableLineItemsFields style=display:none>
          <div class=meta>Rows loop over the <b>line_items</b> sheet (linked by <b>ref_id</b>). Put the loop instructions in the Prompt field above.</div>
        </div>
        <div class=msg id=upmsg></div>
      </div>
    </div>
    <div id=p-run style=display:none>
      <div class="card run-shell">
        <div class=row style="margin-top:0">
          <div style=flex:0.5><label>Execution mode</label><select id=execMode><option value=parallel>parallel</option><option value=sequential>sequential</option></select></div>
          <div style=flex:0.5><label>After action wait (ms)</label><input id=waitMs type=number min=0 step=100 placeholder="0"></div>
        </div>
        <div id=runTabs class=run-tabbar></div>
        <div id=runEmpty class=run-empty>Run a recording from the sidebar to create an output tab.</div>
        <div id=runPane style=display:none>
          <div class=run-title><div><b id=runName>—</b> <span class=meta id=runMeta></span></div></div>
          <div class=actions id=reportActions style=display:none>
            <a id=openReportBtn class="btn" href="#" target="_blank" rel="noopener noreferrer">Open report</a>
            <a id=downloadReportBtn class="btn ghost" href="#" download>Download report</a>
          </div>
          <div class=meta-line id=reportMeta></div>
          <div class=msg id=runmsg></div>
          <pre id=out>select a recording on the left and press Run</pre>
        </div>
      </div>
    </div>
  </div>
</div>
<script>
let CFG={};
let ALL_SCRIPTS=[];
let SUITE=[];let SUITE_MODE='sequential';let SUITE_WAIT=0;
let RUN_TABS=[];let ACTIVE_RUN_TAB_ID='';
function inSuite(n){return SUITE.includes(n)}
function toggleSuite(n,checked){if(checked){if(!SUITE.includes(n))SUITE.push(n)}else{SUITE=SUITE.filter(x=>x!==n)}renderSuite();renderScripts()}
function moveSuite(i,d){const j=i+d;if(j<0||j>=SUITE.length)return;const t=SUITE[i];SUITE[i]=SUITE[j];SUITE[j]=t;renderSuite()}
function removeSuite(n){SUITE=SUITE.filter(x=>x!==n);renderSuite();renderScripts()}
function clearSuite(){SUITE=[];renderSuite();renderScripts()}
function renderSuite(){
  const el=document.getElementById('suiteTray');
  if(!SUITE.length){el.style.display='none';el.innerHTML='';return}
  el.style.display='';
  el.innerHTML=`<div class="card suite-card">
    <div class=suite-head><b>Suite (${SUITE.length})</b>
      <button class=ghost onclick=clearSuite()>Clear</button></div>
    <ol class=suite-list>${SUITE.map((n,i)=>`<li class=suite-item>
      <span class=suite-name>${n}</span>
      <span class=suite-actions><button class=ghost onclick="moveSuite(${i},-1)">↑</button>
      <button class=ghost onclick="moveSuite(${i},1)">↓</button>
      <button class=ghost onclick="removeSuite('${n}')">✕</button></span></li>`).join('')}</ol>
    <div class=row style=margin-bottom:8px>
      <div style=flex:0.6><label style=margin-top:0>Mode</label>
        <select id=suiteMode onchange="SUITE_MODE=this.value">
          <option value=sequential ${SUITE_MODE=='sequential'?'selected':''}>sequential</option>
          <option value=parallel ${SUITE_MODE=='parallel'?'selected':''}>parallel</option></select></div>
      <div style=flex:0.4><label style=margin-top:0>Wait ms</label>
        <input id=suiteWaitMs type=number min=0 step=100 value="${SUITE_WAIT}" oninput="SUITE_WAIT=Number(this.value||0)"></div></div>
    <button onclick=runSuite()>Run suite</button>
    <div class=meta style=margin-top:8px;line-height:1.5">sequential = flow context chains across recordings (later steps see earlier extract/ai_extract outputs)</div>
  </div>`;
}
const EDITORS=[];
function countLines(text){return String(text||'').split('\\n').length}
function setParamsStatus(message,tone){
  const pill=document.getElementById('paramsStatus');
  if(!pill)return;
  pill.textContent=message||'';
  pill.style.display=message?'inline-block':'none';
  pill.style.color=tone==='bad'?'var(--bad)':'var(--mut)';
}
function buildEditorChrome(ta,height,opts){
  const shell=document.createElement('div');
  shell.className='editor-shell plain';
  const head=document.createElement('div');
  head.className='editor-head';
  head.innerHTML=`<div class=editor-head-left><span class=editor-kind>${opts.title}</span>${opts.hint?`<span class=editor-hint>${opts.hint}</span>`:''}</div><div class=editor-head-right>${opts.actionsHtml||''}${opts.json?'<span id="paramsStatus" class="editor-json-status"></span>':''}<span class=editor-status></span></div>`;
  const status=head.querySelector('.editor-status');
  const pane=document.createElement('div');
  pane.className='editor-pane';
  const gutter=document.createElement('div');
  gutter.className='editor-gutter';
  const parent=ta.parentNode;
  parent.insertBefore(shell,ta);
  shell.appendChild(head);
  shell.appendChild(pane);
  pane.appendChild(gutter);
  pane.appendChild(ta);
  ta.classList.add('code-fallback');
  ta.spellcheck=false;
  ta.wrap='off';
  ta.style.height=height+'px';
  return {shell,status,gutter,textarea:ta};
}
function orderedSelectionRange(range){
  return CodeMirror.cmpPos(range.anchor,range.head) <= 0 ? {anchor:range.anchor,head:range.head,from:range.anchor,to:range.head} : {anchor:range.anchor,head:range.head,from:range.head,to:range.anchor};
}
function selectNextOccurrence(cm){
  let query=cm.getSelection();
  if(!query){
    const word=cm.findWordAt(cm.getCursor());
    if(CodeMirror.cmpPos(word.anchor,word.head)===0)return;
    cm.setSelection(word.anchor,word.head);
    query=cm.getSelection();
    if(!query)return;
  }
  const ranges=cm.listSelections().map(orderedSelectionRange).sort((a,b)=>CodeMirror.cmpPos(a.from,b.from));
  const last=ranges[ranges.length-1];
  let cursor=cm.getSearchCursor(query,last.to);
  let found=cursor.findNext();
  if(!found){
    cursor=cm.getSearchCursor(query,CodeMirror.Pos(0,0));
    found=cursor.findNext();
  }
  if(!found)return;
  const next={anchor:cursor.from(),head:cursor.to()};
  const duplicate=ranges.some(range=>CodeMirror.cmpPos(range.from,next.anchor)===0&&CodeMirror.cmpPos(range.to,next.head)===0);
  if(duplicate)return;
  cm.setSelections(cm.listSelections().concat([next]));
  cm.scrollIntoView({from:next.anchor,to:next.head},60);
}
function renderLineNumbers(gutter,text){
  const total=Math.max(1,countLines(text));
  let html='';
  for(let i=1;i<=total;i+=1)html+=`<span>${i}</span>`;
  gutter.innerHTML=html;
}
function makeEditor(id,mode,height,opts={}){
  const ta=document.getElementById(id);
  const chrome=buildEditorChrome(ta,height,opts);
  const state={cm:null,onChange:null};
  function currentValue(){return state.cm?state.cm.getValue():ta.value}
  function updateMeta(){
    const value=currentValue();
    chrome.status.textContent=`${countLines(value)} lines`;
    if(opts.json){
      if(!String(value).trim()){
        chrome.shell.classList.remove('valid','invalid');
        setParamsStatus('', '');
        return;
      }
      try{
        JSON.parse(value);
        chrome.shell.classList.add('valid');
        chrome.shell.classList.remove('invalid');
        setParamsStatus('', '');
      }catch(e){
        chrome.shell.classList.add('invalid');
        chrome.shell.classList.remove('valid');
        setParamsStatus('invalid JSON: '+e.message,'bad');
      }
    }
  }
  function notifyChange(){
    updateMeta();
    if(state.onChange)state.onChange(currentValue());
  }
  if(window.CodeMirror){
    chrome.shell.classList.remove('plain');
    chrome.shell.classList.add('cm-enhanced');
    const cm=CodeMirror.fromTextArea(ta,{mode,theme:'dracula',lineNumbers:true,lineWrapping:false,matchBrackets:true,autoCloseBrackets:true,styleActiveLine:true,showCursorWhenSelecting:true,tabSize:2,indentUnit:2,extraKeys:{'Tab':cm=>cm.execCommand('insertSoftTab'),'Shift-Tab':'indentLess','Cmd-D':selectNextOccurrence,'Ctrl-D':selectNextOccurrence}});
    cm.setSize('100%',height);
    cm.on('change',notifyChange);
    state.cm=cm;
  }else{
    renderLineNumbers(chrome.gutter,ta.value);
    ta.addEventListener('input',()=>{renderLineNumbers(chrome.gutter,ta.value);notifyChange()});
    ta.addEventListener('scroll',()=>{chrome.gutter.scrollTop=ta.scrollTop});
    ta.addEventListener('keydown',e=>{
      if(e.key!=='Tab')return;
      e.preventDefault();
      const start=ta.selectionStart||0;
      const end=ta.selectionEnd||0;
      ta.setRangeText('  ',start,end,'end');
      ta.dispatchEvent(new Event('input'));
    });
  }
  const api={
    get value(){return currentValue()},
    set value(v){
      const next=v==null?'':String(v);
      if(state.cm)state.cm.setValue(next);
      else{
        ta.value=next;
        renderLineNumbers(chrome.gutter,next);
      }
      updateMeta();
    },
    refresh(){if(state.cm)state.cm.refresh()},
    focus(){if(state.cm)state.cm.focus();else ta.focus()},
    setOnChange(fn){state.onChange=fn},
  };
  EDITORS.push(api);
  updateMeta();
  return api;
}
function refreshEditors(){EDITORS.forEach(editor=>setTimeout(()=>editor.refresh(),0))}
function formatParams(){
  try{
    const parsed=JSON.parse(paramsInput.value);
    paramsInput.value=JSON.stringify(parsed,null,2);
    setParamsStatus('', '');
  }catch(e){
    setParamsStatus('invalid JSON: '+e.message,'bad');
  }
}
const recordingNameInput=document.getElementById('name');
const scriptInput=makeEditor('py','python',620,{title:'Python Script',hint:''});
const paramsInput=makeEditor('params',{name:'javascript',json:true},620,{title:'Runtime JSON',hint:'',json:true,actionsHtml:'<div class="editor-icon-actions"><button id="downloadParamsBtn" class="ghost editor-icon-btn" type=button onclick="downloadParamsXlsx()" title="Download XLSX" aria-label="Download XLSX">↓</button><button class="ghost editor-icon-btn" type=button onclick="formatParams()" title="Format JSON" aria-label="Format JSON">{}</button></div>'});
const promptInput=document.getElementById('prompt');
const repeatableLineItemsEnabledInput=document.getElementById('repeatableLineItemsEnabled');
const repeatableLineItemsFields=document.getElementById('repeatableLineItemsFields');
const uploadMessage=document.getElementById('upmsg');
const overwriteSelect=document.getElementById('ov');
const runNameLabel=document.getElementById('runName');
const runMetaLabel=document.getElementById('runMeta');
const runMessage=document.getElementById('runmsg');
const runOutput=document.getElementById('out');
const executionModeSelect=document.getElementById('execMode');
const waitMsInput=document.getElementById('waitMs');
const reportActions=document.getElementById('reportActions');
const openReportBtn=document.getElementById('openReportBtn');
const downloadReportBtn=document.getElementById('downloadReportBtn');
const reportMeta=document.getElementById('reportMeta');
const runTabsTray=document.getElementById('runTabs');
const runPane=document.getElementById('runPane');
const runEmpty=document.getElementById('runEmpty');
const runTabCount=document.getElementById('runTabCount');
const scriptSearchInput=document.getElementById('scriptSearch');
const downloadParamsBtn=document.getElementById('downloadParamsBtn');
async function j(u,o){const r=await fetch(u,o);const t=await r.text();let d;try{d=JSON.parse(t)}catch(e){throw new Error(t)}if(!r.ok)throw new Error(d.detail||t);return d}
function tab(n){
  for(const x of['edit','run'])document.getElementById('p-'+x).style.display=x==n?'':'none';
  document.getElementById('side-edit').classList.toggle('on',n==='edit');
  document.getElementById('side-run').classList.toggle('on',n==='run');
  if(n==='edit')refreshEditors();
  if(n==='run')renderRunTabs();
}
function resetReportActions(){reportActions.style.display='none';openReportBtn.href='#';downloadReportBtn.href='#';downloadReportBtn.removeAttribute('download');reportMeta.textContent=''}
function escapeHtml(text){return String(text==null?'':text).replace(/[&<>"']/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]))}
function getActiveRunTab(){return RUN_TABS.find(tab=>tab.id===ACTIVE_RUN_TAB_ID)||null}
function ensureRunTab(id,label){
  let runTab=RUN_TABS.find(tab=>tab.id===id);
  if(!runTab){
    runTab={id,label,message:'',tone:'ok',running:false,meta:'',output:'',reportUrl:'',reportDownload:'',reportMeta:''};
    RUN_TABS=[runTab].concat(RUN_TABS);
  }else runTab.label=label;
  ACTIVE_RUN_TAB_ID=id;
  renderRunTabs();
  return runTab;
}
function updateRunTab(id,patch){
  const runTab=RUN_TABS.find(tab=>tab.id===id);
  if(!runTab)return;
  Object.assign(runTab,patch);
  renderRunTabs();
}
function selectRunTab(id){ACTIVE_RUN_TAB_ID=id;tab('run');renderRunTabs()}
function renderRunTabs(){
  runTabCount.textContent=String(RUN_TABS.length);
  runTabsTray.innerHTML='';
  RUN_TABS.forEach(tabState=>{
    const btn=document.createElement('button');
    btn.type='button';
    btn.className='run-tab'+(tabState.id===ACTIVE_RUN_TAB_ID?' on':'');
    btn.textContent=tabState.label;
    btn.onclick=()=>selectRunTab(tabState.id);
    runTabsTray.appendChild(btn);
  });
  const active=getActiveRunTab();
  if(!active){
    runEmpty.style.display='';
    runPane.style.display='none';
    runNameLabel.textContent='—';
    runMetaLabel.textContent='';
    runMessage.className='msg';
    runMessage.textContent='';
    runOutput.textContent='';
    resetReportActions();
    return;
  }
  runEmpty.style.display='none';
  runPane.style.display='';
  runNameLabel.textContent=active.label;
  runMetaLabel.textContent=active.meta||'';
  if(active.message){
    runMessage.className='msg '+(active.tone||'ok');
    runMessage.innerHTML=active.running?`${escapeHtml(active.message)} <span class=spin></span>`:escapeHtml(active.message);
  }else{
    runMessage.className='msg';
    runMessage.textContent='';
  }
  runOutput.textContent=active.output||'';
  if(active.reportUrl){
    reportActions.style.display='flex';
    openReportBtn.href=active.reportUrl;
    downloadReportBtn.href=active.reportUrl;
    if(active.reportDownload)downloadReportBtn.setAttribute('download',active.reportDownload);else downloadReportBtn.removeAttribute('download');
    reportMeta.textContent=active.reportMeta||'';
  }else resetReportActions();
}
function toggleRepeatableLineItems(){repeatableLineItemsFields.style.display=repeatableLineItemsEnabledInput.checked?'':'none'}
function buildRepeatableLineItemsPayload(){
  if(!repeatableLineItemsEnabledInput.checked)return null;
  // Sheet name and the ref_id join column are fixed conventions; the loop instructions come from the Prompt field.
  return {enabled:true,sheet_name:'line_items',prompt:(promptInput.value||'').trim()};
}
function applyRecordingConfig(config){
  const prompt=(config&&typeof config.prompt=='string')?config.prompt:'';
  const repeatableBlocks=(config&&Array.isArray(config.repeatable_blocks))?config.repeatable_blocks:[];
  const repeatable=repeatableBlocks.length&&typeof repeatableBlocks[0]=='object'?repeatableBlocks[0]:((config&&config.repeatable_line_items&&typeof config.repeatable_line_items=='object')?config.repeatable_line_items:null);
  promptInput.value=prompt||((repeatable&&repeatable.prompt)||'');
  repeatableLineItemsEnabledInput.checked=!!(repeatable&&repeatable.enabled!==false);
  toggleRepeatableLineItems();
}
async function loadCfg(){CFG=await j('/api/config');document.getElementById('cfg').textContent=`bucket: ${CFG.bucket} · pg: ${CFG.pg.host}:${CFG.pg.port}/${CFG.pg.db} · agent: local`;waitMsInput.value=String(CFG.default_after_action_wait_ms ?? 0)}
function renderScripts(){
  const query=(scriptSearchInput.value||'').trim().toLowerCase();
  const visible=ALL_SCRIPTS.filter(s=>!query||s.name.toLowerCase().includes(query));
  document.getElementById('count').textContent=`${visible.length}/${ALL_SCRIPTS.length} in ${CFG.bucket||''}`.trim();
  document.getElementById('list').innerHTML=visible.map(s=>`<div class="script ${inSuite(s.name)?'sel':''}">
   <div class=script-main><div class=script-top>
   <input type=checkbox class=suitechk title="add to suite" ${inSuite(s.name)?'checked':''} onchange="toggleSuite('${s.name}',this.checked)">
   <div class=script-copy><div class=nm>${s.name}</div>
   <div class=sub>${s.params_key?s.params_key.split('/').pop():'(no params)'} · <span class="badge ${s.has_db?'db':'nodb'}">${s.has_db?'DB ✓':'no DB'}</span></div></div></div></div>
   <div class=script-actions><button class=script-btn onclick="runIt('${s.name}')">Run</button><button class="script-btn ghost" onclick="loadOne('${s.name}')">Edit</button></div></div>`).join('')||'<div class=meta>no recordings</div>'}
async function loadScripts(){const d=await j('/api/scripts');ALL_SCRIPTS=d.scripts||[];if(!CFG.bucket)CFG.bucket=d.bucket||'';renderScripts();renderSuite()}
async function loadOne(n){const d=await j('/api/script?name='+encodeURIComponent(n));recordingNameInput.value=d.name;scriptInput.value=d.py_text;
  const payload={params:d.params.length?d.params:[{}]};const lineItems=Array.isArray(d.line_items)?d.line_items:(Array.isArray(d.multi_line)?d.multi_line:[]);if(lineItems.length)payload.line_items=lineItems;
  paramsInput.value=JSON.stringify(payload,null,2);applyRecordingConfig(d.recording_config||{});tab('edit');
  uploadMessage.className='msg ok';uploadMessage.textContent=`loaded ${d.name} · placeholders: ${d.placeholders.join(', ')||'none'} · DB: ${d.db?'yes':'no'}`}
function clearForm(){recordingNameInput.value=scriptInput.value=paramsInput.value=promptInput.value='';repeatableLineItemsEnabledInput.checked=false;toggleRepeatableLineItems();uploadMessage.className='msg'}
async function doUpload(){const b=document.getElementById('upbtn');b.disabled=true;uploadMessage.className='msg';uploadMessage.textContent='';
  try{let p;try{p=JSON.parse(paramsInput.value)}catch(e){setParamsStatus('invalid JSON: '+e.message,'bad');paramsInput.focus();return}
   const d=await j('/api/upload',{method:'POST',headers:{'content-type':'application/json'},body:JSON.stringify({name:recordingNameInput.value,script:scriptInput.value,params:p,fmt:'xlsx',overwrite:overwriteSelect.value=='true',prompt:promptInput.value,repeatable_blocks:(()=>{const block=buildRepeatableLineItemsPayload();return block?[block]:null;})()})});
   let m=`uploaded → ${d.py_key} + ${d.params_key.split('/').pop()} · start_url=${d.start_url||'?'} · DB ${d.db.inserted?'inserted':(d.db.conflict?'conflict':'updated')} (id ${d.db.id||'-'})`;
   if(d.recording_config_key)m+=` · config: ${d.recording_config_key.split('/').pop()}`;
   if(d.db_error)m+=` · ⚠ DB: ${d.db_error}`;if(d.missing_placeholders.length)m+=` · ⚠ missing params for: ${d.missing_placeholders.join(', ')}`;
   uploadMessage.className='msg '+(d.db_error?'err':'ok');uploadMessage.textContent=m;loadScripts()}
  catch(e){uploadMessage.className='msg err';uploadMessage.textContent=e.message}finally{b.disabled=false}}
function safeName(v){return (v||'').trim().replace(/[^A-Za-z0-9._-]+/g,'_').replace(/^[._]+|[._]+$/g,'')}
async function downloadParamsXlsx(){
  let payload;
  try{payload=JSON.parse(paramsInput.value)}catch(e){setParamsStatus('invalid JSON: '+e.message,'bad');paramsInput.focus();return}
  const original=downloadParamsBtn.textContent;
  downloadParamsBtn.disabled=true;
  downloadParamsBtn.textContent='Preparing...';
  try{
    const d=await j('/api/params-xlsx',{method:'POST',headers:{'content-type':'application/json'},body:JSON.stringify({name:recordingNameInput.value||'recording_name',params:payload})});
    const a=document.createElement('a');
    a.href=d.download_url;
    a.download=d.download_name||'params_preview.xlsx';
    document.body.appendChild(a);
    a.click();
    a.remove();
    uploadMessage.className='msg ok';
    uploadMessage.textContent=`download ready → ${d.download_name}`;
  }catch(e){
    uploadMessage.className='msg err';
    uploadMessage.textContent=e.message;
  }finally{
    downloadParamsBtn.disabled=false;
    downloadParamsBtn.textContent=original;
  }
}
async function runIt(n){
  const tabId=`recording:${n}`;
  ensureRunTab(tabId,n);
  tab('run');
  resetReportActions();
  const waitValue=Number(waitMsInput.value||0);
  updateRunTab(tabId,{
    message:'running (uses your local agent; may take minutes)',
    tone:'ok',
    running:true,
    meta:`mode=${executionModeSelect.value} · wait=${waitValue}ms`,
    output:'$ aetherion agent "ACT Agent" ... --wait\\n(waiting for your local worker)',
    reportUrl:'',
    reportDownload:'',
    reportMeta:''
  });
  try{const body={name:n,execution_mode:executionModeSelect.value,after_action_wait_ms:waitValue};
   if(safeName(recordingNameInput.value)===n&&paramsInput.value.trim()){let p;try{p=JSON.parse(paramsInput.value)}catch(e){throw new Error('params JSON invalid: '+e.message)};body.parameters=p}
   const d=await j('/api/run',{method:'POST',headers:{'content-type':'application/json'},body:JSON.stringify(body)});
   let m=d.ok?'completed (exit 0)':'failed (exit '+d.returncode+')';
   if(d.prepared_recording_count&&d.prepared_recording_count>1)m+=` · expanded to ${d.prepared_recording_count} prepared runs`;
   if(d.inline_parameter_keys&&d.inline_parameter_keys.length)m+=` · inline params: ${d.inline_parameter_keys.join(', ')}`;
   if(d.report_local)m+=' · report: '+d.report_local;
   updateRunTab(tabId,{
     message:m,
     tone:d.ok?'ok':'err',
     running:false,
     meta:`mode=${d.execution_mode} · wait=${d.used_after_action_wait_ms}ms`,
     output:(d.stdout||'')+(d.stderr?'\\n--- stderr ---\\n'+d.stderr:''),
     reportUrl:d.report_url||'',
     reportDownload:d.report_url?d.report_url.split('/').pop():'',
     reportMeta:d.report_local||d.report_key||''
   })}
  catch(e){updateRunTab(tabId,{message:e.message,tone:'err',running:false,output:e.message})}}
async function runSuite(){
  if(!SUITE.length)return;
  const label='Suite: '+SUITE.join(' → ');
  const tabId='suite:'+SUITE.join('|');
  ensureRunTab(tabId,label);
  tab('run');
  resetReportActions();
  updateRunTab(tabId,{
    message:'running suite (uses your local agent; may take minutes)',
    tone:'ok',
    running:true,
    meta:`mode=${SUITE_MODE} · ${SUITE.length} recordings · wait=${SUITE_WAIT}ms`,
    output:'$ aetherion agent "ACT Agent" ... --wait\\n(waiting for your local worker)',
    reportUrl:'',
    reportDownload:'',
    reportMeta:''
  });
  try{const body={recordings:SUITE.map(n=>({name:n})),execution_mode:SUITE_MODE,after_action_wait_ms:SUITE_WAIT};
   const d=await j('/api/run-suite',{method:'POST',headers:{'content-type':'application/json'},body:JSON.stringify(body)});
   let m=d.ok?'suite completed (exit 0)':'suite failed (exit '+d.returncode+')';
   m+=` · ${(d.recordings||[]).length} recordings`;
   if(d.report_local)m+=' · report: '+d.report_local;
   updateRunTab(tabId,{
     message:m,
     tone:d.ok?'ok':'err',
     running:false,
     meta:`suite=${d.suite_id} · mode=${d.execution_mode} · wait=${d.used_after_action_wait_ms}ms`,
     output:(d.stdout||'')+(d.stderr?'\\n--- stderr ---\\n'+d.stderr:''),
     reportUrl:d.report_url||'',
     reportDownload:d.report_url?d.report_url.split('/').pop():'',
     reportMeta:d.report_local||d.report_key||''
   })}
  catch(e){updateRunTab(tabId,{message:e.message,tone:'err',running:false,output:e.message})}}
scriptSearchInput.addEventListener('input', renderScripts);
toggleRepeatableLineItems();
renderRunTabs();
tab('edit');
loadCfg();loadScripts();
</script></body></html>"""


if __name__ == "__main__":
    host = "127.0.0.1"
    port = int(os.environ.get("PORT", "8765"))
    print(
        f"agent_shubham → http://localhost:{port}   "
        f"(bucket={BUCKET}, pg={PG['host']}:{PG['port']}/{PG['dbname']})"
    )
    uvicorn.run(app, host=host, port=port, log_level="info")
